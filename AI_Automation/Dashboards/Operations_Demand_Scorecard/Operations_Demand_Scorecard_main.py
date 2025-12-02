from datetime import datetime, timedelta
import pandas as pd
import os
from Unleashed_Data.Unleashed_Load_Parralelize import get_data_parallel
from AI_Automation.Dashboards.Operations_Demand_Scorecard.Operations_Demand_Scorecard_helper import (unwrap_sales_orders,
                                                                                                     unwrap_warehouse_sales_orders,
                                                                                                     get_parts_list, clean_sales_orders, last_modified_on)


#custom statuses are grouped into parked

def first_try_dashboard():
    print("First Try Dashboard")

    today = datetime.today()
    yesterday = today - timedelta(days=1)
    one_week_ago = today - timedelta(days=7)
    far_back = datetime(2020, 1, 1)

    #get open orders from (one year and week ago) to (week ago) this finds all open orders we had before reporting week
    start_date_far_back = far_back.strftime('%Y-%m-%d')
    # end_date_one_week_ago = one_week_ago.strftime('%Y-%m-%d')
    end_date_one_week_ago = '2025-11-15'
    df_salesOrders_before_week = get_data_parallel(unleashed_data_name="SalesOrdersDate", start_date=start_date_far_back,  end_date=end_date_one_week_ago)

    df_salesOrders_before_week["ui_status"] = df_salesOrders_before_week["CustomOrderStatus"].fillna(df_salesOrders_before_week["OrderStatus"])
    print(start_date_far_back)
    print(end_date_one_week_ago)

    # get old open orders
    old_status_counts = df_salesOrders_before_week['ui_status'].value_counts()
    old_completed = old_status_counts.get('Completed', 0)
    old_deleted = old_status_counts.get('Deleted', 0)
    old_parked = old_status_counts.get('Parked', 0)
    old_backordered = old_status_counts.get('Backordered', 0)
    old_placed = old_status_counts.get('Placed', 0)
    old_hold = old_status_counts.get('HOLD', 0)
    old_accounting = old_status_counts.get('Accounting', 0)
    old_ready_to_ship = old_status_counts.get('Ready to Ship', 0)

    # old_open_orders = old_placed + old_parked + old_backordered
    old_open_orders = len(df_salesOrders_before_week) - old_completed - old_deleted

    old_open_other = old_open_orders - old_parked - old_backordered - old_placed - old_hold - old_accounting - old_ready_to_ship



    #get orders during this week so we can analyze orders we got this week
    # start_date_one_week_ago = one_week_ago.strftime('%Y-%m-%d')
    start_date_one_week_ago = end_date_one_week_ago
    # end_date_today = today.strftime('%Y-%m-%d') #end date gives data up to yesterday since end date is not inclusive
    end_date_today = '2025-11-22'
    df_salesOrders_week = get_data_parallel(unleashed_data_name="SalesOrdersDate", start_date=start_date_one_week_ago,  end_date=end_date_today)

    df_salesOrders_week["ui_status"] = df_salesOrders_week["CustomOrderStatus"].fillna(df_salesOrders_week["OrderStatus"])

    print(start_date_one_week_ago)
    print(end_date_today)

    # Order status counts
    new_num_orders = len(df_salesOrders_week)
    status_counts = df_salesOrders_week['ui_status'].value_counts()
    this_week_completed = status_counts.get('Completed', 0)
    new_deleted = status_counts.get('Deleted', 0)
    new_completed = status_counts.get('Completed', 0)
    new_parked = status_counts.get('Parked', 0)
    new_backordered = status_counts.get('Backordered', 0)
    new_placed = status_counts.get('Placed', 0)
    new_open_orders = new_placed + new_parked + new_backordered

    new_hold = status_counts.get('HOLD', 0)
    new_accounting = status_counts.get('Accounting', 0)
    new_ready_to_ship = status_counts.get('Ready to Ship', 0)

    new_open_orders = new_num_orders - new_completed - new_deleted

    new_open_other = new_open_orders - new_parked - new_backordered - new_placed - new_hold - new_accounting - new_ready_to_ship
    print("New open orders: ", new_open_orders)



    ops_summary_this_week = pd.DataFrame({
        "Metric": [
            "Old Open Orders",
            "New Orders This Week",
            "TOTAL ORDERS",
            "New Orders This Week Completed",
            "New Orders This Week Deleted",
            "TOTAL OPEN ORDERS"
        ],
        "Value": [
            old_open_orders,
            new_num_orders,
            old_open_orders + new_num_orders,
            this_week_completed,
            new_deleted,
            old_open_orders + new_open_orders
        ]
    })
    print(ops_summary_this_week)



    print("///////////////////////")
    ops_summary_open_orders = pd.DataFrame({
        "Metric": [
            "Old Open Orders Parked",
            "Old Open Orders Backordered",
            "Old Open Orders Placed",
            "Old Open Orders Accounting",
            "Old Open Orders Ready to Ship",
            "Old Open Orders Hold",
            "Old Open Orders Other",
            "TOTAL OLD OPEN ORDERS",
            "New Open Orders Parked",
            "New Open Orders Backordered",
            "New Open Orders Placed",
            "New Open Orders Accounting",
            "New Open Orders Ready to Ship",
            "New Open Orders Hold",
            "New Open Orders Other",
            "TOTAL NEW OPEN ORDERS",
            "TOTAL OPEN ORDERS"
        ],
        "Value": [
            old_parked,
            old_backordered,
            old_placed,
            old_accounting,
            old_ready_to_ship,
            old_hold,
            old_open_other,
            old_open_orders,
            new_parked,
            new_backordered,
            new_placed,
            new_accounting,
            new_ready_to_ship,
            new_hold,
            new_open_other,
            new_open_orders,
            old_open_orders + new_open_orders
        ]
    })
    print(ops_summary_open_orders)




    print("///////////////////////")
    df_SalesOrders_completed_dates = get_data_parallel(unleashed_data_name="SalesOrders", start_date=start_date_one_week_ago, end_date=end_date_today)  # Sales Orders By Completed Date
    total_completed_orders_this_week = df_SalesOrders_completed_dates['OrderNumber'].nunique()
    ops_summary_completed = pd.DataFrame({
        "Metric": [
            "Old Orders Completed",
            "New Orders Completed",
            "TOTAL COMPLETED ORDERS"
        ],
        "Value": [
            total_completed_orders_this_week - this_week_completed,
            this_week_completed,
            total_completed_orders_this_week,
        ]
    })
    print(ops_summary_completed)






    #Now start looking at
    #completed bike orders = Completed orders - new orders completed

    #of all the completed orders, count the bike orders
    print("///////////////////////")
    df_unwrapped_salesorders_completed_date = unwrap_warehouse_sales_orders(df_SalesOrders_completed_dates)
    df_unwrapped_salesorders_JamN_bikes_completed_dates = df_unwrapped_salesorders_completed_date.loc[
        (df_unwrapped_salesorders_completed_date['WarehouseName'] == 'Jam-N Logistics') &
        (df_unwrapped_salesorders_completed_date['ProductGroup'] == 'Bikes')] #DO I NEED TO EXCLUDE CUSTOMER TYPE = COSTCO? That is what makes the orderquantity sum large
    Total_num_Bike_Orders_JamN = df_unwrapped_salesorders_JamN_bikes_completed_dates['OrderNumber'].nunique()
    num_bikes_fulfilled_JamN = df_unwrapped_salesorders_JamN_bikes_completed_dates['OrderQuantity'].sum()

    ops_bike_summary_completed = pd.DataFrame({
        "Metric": [
            "Total # of Bike Orders: Jam-N",
            "# of Bikes Fulfilled: Jam-N",
            "Jam-N Bikes Per Order"
        ],
        "Value": [
            Total_num_Bike_Orders_JamN,
            num_bikes_fulfilled_JamN,
            num_bikes_fulfilled_JamN / Total_num_Bike_Orders_JamN
        ]
    })
    print(ops_bike_summary_completed)





    #Accessories
    print("///////////////////////")
    df_unwrapped_salesorders_accessories_order_dates = df_unwrapped_salesorders_completed_date.loc[df_unwrapped_salesorders_completed_date['ProductGroup'] == 'Accessories']
    total_accessories_orders = df_unwrapped_salesorders_accessories_order_dates['OrderNumber'].nunique()
    accessory_units_fulfilled = df_unwrapped_salesorders_accessories_order_dates['OrderQuantity'].sum()

    ops_accessory_summary_completed = pd.DataFrame({
        "Metric": [
            "Total Accessories Orders",
            "Accessory Units Fulfilled",
            "Accessories Per Completed Order"
        ],
        "Value": [
            total_accessories_orders,
            accessory_units_fulfilled,
            accessory_units_fulfilled / total_accessories_orders
        ]
    })
    print(ops_accessory_summary_completed)







    #Parts
    print("///////////////////////")
    df_unwrapped_salesorders_parts_completed_dates = df_unwrapped_salesorders_completed_date.loc[df_unwrapped_salesorders_completed_date['ProductGroup'].isin(get_parts_list())]
    total_parts_orders = df_unwrapped_salesorders_parts_completed_dates['OrderNumber'].nunique()
    print("Total Parts Orders: ", total_parts_orders)
    parts_units_fulfilled = df_unwrapped_salesorders_parts_completed_dates['OrderQuantity'].sum()
    print("Parts Units Fulfilled: ", parts_units_fulfilled)

    ops_parts_summary_completed = pd.DataFrame({
        "Metric": [
            "Total Parts Orders",
            "Parts Units Fulfilled",
            "Parts per Order"
        ],
        "Value": [
            total_parts_orders,
            parts_units_fulfilled,
            parts_units_fulfilled / total_parts_orders
        ]
    })
    print(ops_parts_summary_completed)


    #MAKE EXCEL FILE
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    # --- Your existing summaries (already defined earlier in your script) ---
    # ops_summary_this_week
    # ops_summary_open_orders
    # ops_summary_completed
    # ops_bike_summary_completed
    # ops_accessory_summary_completed
    # ops_parts_summary_completed

    # Collect all summaries with titles
    sections = [
        ("This Week Summary", ops_summary_this_week),
        ("Open Orders Summary", ops_summary_open_orders),
        ("Completed Orders Summary", ops_summary_completed),
        ("Bike Orders Summary", ops_bike_summary_completed),
        ("Accessories Summary", ops_accessory_summary_completed),
        ("Parts Summary", ops_parts_summary_completed),
    ]

    # Build a combined DataFrame with titles + spacing
    combined = []
    for title, df in sections:
        # Title row
        combined.append(pd.DataFrame({"Metric": [title], "Value": [""]}))
        # Actual section
        combined.append(df)
        # Blank spacer row
        combined.append(pd.DataFrame({"Metric": [""], "Value": [""]}))

    final_df = pd.concat(combined, ignore_index=True)

    # Save to Excel at the requested path
    file_path = fr"C:\Users\joshu\Documents\Ops_Dashboard_Draft1.xlsx"
    final_df.to_excel(file_path, index=False)

    # Optional: format titles bold
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):  # only "Metric" column
        cell = row[0]
        if cell.value in [title for title, _ in sections]:
            cell.font = Font(bold=True)
    wb.save(file_path)


def second_try_dashboard():
    print("Second Try Dashboard")
    today = datetime.today()
    yesterday = today - timedelta(days=1)
    one_week_ago = today - timedelta(days=7)
    far_back = datetime(2025, 1, 1)

    # get open orders from (one year and week ago) to (week ago) this finds all open orders we had before reporting week
    start_date_far_back = far_back.strftime('%Y-%m-%d')
    end_date_one_week_ago = one_week_ago.strftime('%Y-%m-%d')
    # end_date_one_week_ago = '2025-11-11'
    df_salesOrders_before_week = get_data_parallel(unleashed_data_name="SalesOrdersDate",
                                                   start_date=start_date_far_back, end_date=end_date_one_week_ago)
    print(start_date_far_back)
    print(end_date_one_week_ago)

    # get old open orders
    old_status_counts = df_salesOrders_before_week['OrderStatus'].value_counts()
    old_parked = old_status_counts.get('Parked', 0)
    old_backordered = old_status_counts.get('Backordered', 0)
    old_placed = old_status_counts.get('Placed', 0)
    old_open_orders = old_placed + old_parked + old_backordered

    # get orders during this week so we can analyze orders we got this week
    start_date_one_week_ago = end_date_one_week_ago
    end_date_today = today.strftime('%Y-%m-%d')  # end date gives data up to yesterday since end date is not inclusive
    # end_date_today = '2025-11-18'
    df_salesOrders_week = get_data_parallel(unleashed_data_name="SalesOrdersDate", start_date=start_date_one_week_ago,
                                            end_date=end_date_today)

    df_salesOrders_week["ui_status"] = df_salesOrders_week["CustomOrderStatus"].fillna(df_salesOrders_week["OrderStatus"])

    print("This thing: ", len(df_salesOrders_week))
    file_path = fr"C:\Users\joshu\Documents\custom_statuses_salesOrders.xlsx"
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    df_salesOrders_week.to_excel(file_path, index=False)
    print(f"Excel file written to: {file_path}")
    print("DID IT")

    print(start_date_one_week_ago)
    print(end_date_today)

    # Order status counts
    new_num_orders = len(df_salesOrders_week)
    status_counts = df_salesOrders_week['OrderStatus'].value_counts()
    this_week_completed = status_counts.get('Completed', 0)
    new_deleted = status_counts.get('Deleted', 0)
    new_parked = status_counts.get('Parked', 0)
    new_backordered = status_counts.get('Backordered', 0)
    new_placed = status_counts.get('Placed', 0)
    new_open_orders = new_placed + new_parked + new_backordered

    df_SalesOrders_completed_dates = get_data_parallel(unleashed_data_name="SalesOrders", start_date=start_date_one_week_ago, end_date=end_date_today)  # Sales Orders By Completed Date
    total_completed_orders_this_week = df_SalesOrders_completed_dates['OrderNumber'].nunique()

    ops_summary_productivity_WTD = pd.DataFrame({
        "Metric": [
            "New Orders",
            "Completed Orders",
            "*Deleted (DELETED ORDERS OF CREATED WTD)",
            "*Completed % (USES DELETED)"
        ],
        "Value": [
            new_num_orders,
            total_completed_orders_this_week,
            new_deleted,
            (total_completed_orders_this_week + new_deleted) / new_num_orders
        ]
    })
    print(ops_summary_productivity_WTD)

    print("///////////////////////")
    df_total_orderDates = pd.concat([df_salesOrders_before_week, df_salesOrders_week], axis=0)
    df_total_orderDates = clean_sales_orders(df_total_orderDates)

    ops_summary_work_in_progres_ytd = pd.DataFrame({
        "Metric": [
            "Parked",
            "Backordered",
            "Hold + Accounting(Sales)",
            "Pre Orders",
            "Placed(Bikes) or Ready to Ship(parts & accessories)",
            "Total Open Orders",
            "# of Units associated with Pre-Orders"
        ],
        "Value": [
            old_parked + new_parked,
            old_backordered + new_backordered,
            "N/A",
            "Zach",
            "Don't Know",
            old_open_orders + new_open_orders,
            "Use once figure out open orders"
        ]
    })
    print(ops_summary_work_in_progres_ytd)
    #
    # file_path = fr"C:\Users\joshu\Documents\all_orders_this week.xlsx"
    # os.makedirs(os.path.dirname(file_path), exist_ok=True)
    # df_total_orderDates.to_excel(file_path, index=False)
    # print(f"Excel file written to: {file_path}")








    # print("///////////////////////")
    # df_salesOrders_week_unwrapped = unwrap_sales_orders(df_salesOrders_week)
    # df_bikeOrders_orderDate_this_week = df_salesOrders_week_unwrapped.loc[
    #     (df_salesOrders_week_unwrapped['WarehouseName'] == 'Jam-N Logistics') &
    #     (df_salesOrders_week_unwrapped[
    #          'ProductGroup'] == 'Bikes')]
    # Total_num_Bike_Orders_JamN_orderDate = df_bikeOrders_orderDate_this_week['OrderNumber'].nunique()



    df_unwrapped_salesorders_completed_date = unwrap_warehouse_sales_orders(df_SalesOrders_completed_dates)
    df_unwrapped_salesorders_JamN_bikes_completed_dates = df_unwrapped_salesorders_completed_date.loc[
        (df_unwrapped_salesorders_completed_date['WarehouseName'] == 'Jam-N Logistics') &
        (df_unwrapped_salesorders_completed_date[
             'ProductGroup'] == 'Bikes')]
    Total_num_Bike_Orders_JamN = df_unwrapped_salesorders_JamN_bikes_completed_dates['OrderNumber'].nunique()
    num_bikes_fulfilled_JamN = df_unwrapped_salesorders_JamN_bikes_completed_dates['OrderQuantity'].sum()

    ops_bike_summary_completed = pd.DataFrame({
        "Metric": [
            "Total # of Bike Orders: Jam-N",
            "# of Bikes Fulfilled: Jam-N",
            "Jam-N Bikes Per Order"
        ],
        "Value": [
            Total_num_Bike_Orders_JamN,
            num_bikes_fulfilled_JamN,
            num_bikes_fulfilled_JamN / Total_num_Bike_Orders_JamN
        ]
    })
    print(ops_bike_summary_completed)




    #Parts
    print("///////////////////////")
    df_unwrapped_salesorders_parts_completed_dates = df_unwrapped_salesorders_completed_date.loc[df_unwrapped_salesorders_completed_date['ProductGroup'].isin(get_parts_list())]
    # df_unwrapped_salesorders_parts_completed_dates = df_unwrapped_salesorders_completed_date.loc[df_unwrapped_salesorders_completed_date['Warehouse'].isin(['Brickyard', ''])]
    total_parts_orders = df_unwrapped_salesorders_parts_completed_dates['OrderNumber'].nunique()
    parts_units_fulfilled = df_unwrapped_salesorders_parts_completed_dates['OrderQuantity'].sum()

    df_unwrapped_salesorders_accessories_order_dates = df_unwrapped_salesorders_completed_date.loc[
        df_unwrapped_salesorders_completed_date['ProductGroup'] == 'Accessories']
    total_accessories_orders = df_unwrapped_salesorders_accessories_order_dates['OrderNumber'].nunique()
    accessory_units_fulfilled = df_unwrapped_salesorders_accessories_order_dates['OrderQuantity'].sum()

    ops_parts_accessories_summary_completed = pd.DataFrame({
        "Metric": [
            "Total Parts Orders Completed",
            "Parts Units Fulfilled",
            "Total Accessories Orders",
            "Accessory Units Fulfilled"
        ],
        "Value": [
            total_parts_orders,
            parts_units_fulfilled,
            total_accessories_orders,
            accessory_units_fulfilled
        ]
    })
    print(ops_parts_accessories_summary_completed)







    #
    #
    # #Inventory
    # df_stockonhand = get_data_parallel(unleashed_data_name="StockOnHand", end_date=today.strftime('%Y-%m-%d'))
    # print("stock columns: ", df_stockonhand.columns)
    #
    # #bikes in stock
    # df_stockonhand_bikes_in_stock = df_stockonhand.loc[(df_stockonhand['ProductGroupName'] == 'Bikes') & (df_stockonhand['Warehouse'].isin(['Brickyard', 'Jam-N Logistics']))]
    # bikes_in_stock = df_stockonhand_bikes_in_stock['QtyOnHand'].sum()
    #
    #
    # #bikes in backorder
    # df_total_orderDates
    # df_unwrapped_salesorders_orderDate_backorder_bikes = df_total_orderDates.loc[(df_total_orderDates['Status'] == 'Backordered') & (df_total_orderDates['ProductGroup'] == 'Bikes')]
    # bikes_in_back_order = df_unwrapped_salesorders_orderDate_backorder_bikes['QtyOnHand'].sum()
    #
    # ops_inventory_health_summary_completed = pd.DataFrame({
    #     "Metric": [
    #         "Bikes in stock",
    #         "CPO Bikes in Stock",
    #         "Bikes in Do Not Sell WH",
    #         "Bikes awaiting shipment",
    #         "Bikes in back order",
    #         "Parts in stock",
    #         "Parts awaiting shipment",
    #         "Accessories in stock",
    #         "Accessories awaiting shipment",
    #         "Accessories in back orders"
    #     ],
    #     "Value": [
    #         bikes_in_stock,
    #         "N/A",
    #         "N/A",
    #         "N/A",
    #         bikes_in_back_order,
    #         "N/A",
    #         "N/A",
    #         "N/A",
    #         "N/A",
    #         "N/A",
    #     ]
    # })
    # print(ops_parts_accessories_summary_completed)











    # MAKE EXCEL FILE
    from openpyxl import load_workbook
    from openpyxl.styles import Font


    # Collect all summaries with titles
    sections = [
        ("Productivity WTD", ops_summary_productivity_WTD),
        ("Work in Progress YTD", ops_summary_work_in_progres_ytd),
        ("Bike Orders", ops_bike_summary_completed),
        ("Parts Orders", ops_parts_accessories_summary_completed)
    ]

    # Build a combined DataFrame with titles + spacing
    combined = []
    for title, df in sections:
        # Title row
        combined.append(pd.DataFrame({"Metric": [title], "Value": [""]}))
        # Actual section
        combined.append(df)
        # Blank spacer row
        combined.append(pd.DataFrame({"Metric": [""], "Value": [""]}))

    final_df = pd.concat(combined, ignore_index=True)

    # Save to Excel at the requested path
    file_path = fr"C:\Users\joshu\Documents\Ops_Dashboard_Draft2.xlsx"
    final_df.to_excel(file_path, index=False)

    # Optional: format titles bold
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):  # only "Metric" column
        cell = row[0]
        if cell.value in [title for title, _ in sections]:
            cell.font = Font(bold=True)
    wb.save(file_path)
















def third_try_dashboard():
    print("Third Try Dashboard")
    today = datetime.today()
    yesterday = today - timedelta(days=1)
    one_week_ago = today - timedelta(days=7)
    far_back = datetime(2020, 1, 1)

    # get open orders from (one year and week ago) to (week ago) this finds all open orders we had before reporting week
    start_date_far_back = far_back.strftime('%Y-%m-%d')
    end_date_one_week_ago = one_week_ago.strftime('%Y-%m-%d')
    end_date_one_week_ago = '2025-11-15'
    df_salesOrders_before_week = get_data_parallel(unleashed_data_name="SalesOrdersDate",
                                                   start_date=start_date_far_back, end_date=end_date_one_week_ago)


    # file_path = fr"C:\Users\joshu\Documents\last_week_orderDates.xlsx"
    # os.makedirs(os.path.dirname(file_path), exist_ok=True)
    # df_salesOrders_before_week.to_excel(file_path, index=False)
    # print(f"Excel file written to: {file_path}")

    df_salesOrders_before_week["ui_status"] = df_salesOrders_before_week["CustomOrderStatus"].fillna(df_salesOrders_before_week["OrderStatus"])
    print(start_date_far_back)
    print(end_date_one_week_ago)

    # get old open orders
    old_status_counts = df_salesOrders_before_week['ui_status'].value_counts()
    old_completed = old_status_counts.get('Completed', 0)
    old_deleted = old_status_counts.get('Deleted', 0)
    old_parked = old_status_counts.get('Parked', 0)
    old_backordered = old_status_counts.get('Backordered', 0)
    old_placed = old_status_counts.get('Placed', 0)
    old_hold = old_status_counts.get('HOLD', 0)
    old_accounting = old_status_counts.get('Accounting', 0)
    old_ready_to_ship = old_status_counts.get('Ready to Ship', 0)
    old_ready_to_ship_STG = old_status_counts.get('Ready Ship STG', 0)

    # old_open_orders = old_placed + old_parked + old_backordered
    old_open_orders = len(df_salesOrders_before_week) - old_completed - old_deleted


    # get orders during this week so we can analyze orders we got this week
    start_date_one_week_ago = end_date_one_week_ago
    # end_date_today = today.strftime('%Y-%m-%d')  # end date gives data up to yesterday since end date is not inclusive
    end_date_today = '2025-11-22'
    df_salesOrders_week = get_data_parallel(unleashed_data_name="SalesOrdersDate", start_date=start_date_one_week_ago,
                                            end_date=end_date_today)

    df_salesOrders_week["ui_status"] = df_salesOrders_week["CustomOrderStatus"].fillna(df_salesOrders_week["OrderStatus"])

    # get number of deleted during the week
    df_salesOrders_before_week_fixed_lmd = last_modified_on(df_salesOrders_before_week)
    df_salesOrders_before_week_fixed_lmd_deleted = df_salesOrders_before_week_fixed_lmd.loc[
        df_salesOrders_before_week_fixed_lmd['OrderStatus'] == 'Deleted']
    start_date_lmd = datetime.strptime(start_date_one_week_ago, "%Y-%m-%d").date()
    end_date_lmd = datetime.strptime(end_date_today, "%Y-%m-%d").date()
    mask = (df_salesOrders_before_week_fixed_lmd_deleted['LastModifiedOn'] >= start_date_lmd) & (
                df_salesOrders_before_week_fixed_lmd_deleted['LastModifiedOn'] < end_date_lmd)
    df_old_deleted = df_salesOrders_before_week_fixed_lmd_deleted.loc[mask]
    print("THING: ", df_old_deleted['OrderNumber'])
    old_deleted_everything = len(df_old_deleted)
    print("THING2: ", old_deleted_everything)


    print(start_date_one_week_ago)
    print(end_date_today)

    # Order status counts
    new_num_orders = len(df_salesOrders_week)
    status_counts = df_salesOrders_week['ui_status'].value_counts()
    this_week_completed = status_counts.get('Completed', 0)
    new_deleted = status_counts.get('Deleted', 0)
    new_completed = status_counts.get('Completed', 0)
    new_parked = status_counts.get('Parked', 0)
    new_backordered = status_counts.get('Backordered', 0)
    new_placed = status_counts.get('Placed', 0)
    new_open_orders = new_placed + new_parked + new_backordered

    new_hold = status_counts.get('HOLD', 0)
    new_accounting = status_counts.get('Accounting', 0)
    new_ready_to_ship = status_counts.get('Ready to Ship', 0)
    new_ready_to_ship_STG = status_counts.get('Ready Ship STG', 0)

    new_open_orders = new_num_orders - new_completed - new_deleted


    df_SalesOrders_completed_dates = get_data_parallel(unleashed_data_name="SalesOrders", start_date=start_date_one_week_ago, end_date=end_date_today)  # Sales Orders By Completed Date
    total_completed_orders_this_week = df_SalesOrders_completed_dates['OrderNumber'].nunique()

    ops_summary_productivity_WTD = pd.DataFrame({
        "Metric": [
            "New Orders",
            "Completed Orders",
            "*Deleted",
            "*Completed %"
        ],
        "Value": [
            new_num_orders,
            total_completed_orders_this_week,
            old_deleted_everything + new_deleted,
            (total_completed_orders_this_week + old_deleted_everything + new_deleted) / new_num_orders
        ]
    })
    print(ops_summary_productivity_WTD)

    print("///////////////////////")

    df_salesOrders_before_week_unwrapped = clean_sales_orders(df_salesOrders_before_week)
    df_salesOrders_week_unwrapped = clean_sales_orders(df_salesOrders_week)



    #before this week
    #bikes
    df_placed_bikes_before_week = df_salesOrders_before_week_unwrapped.loc[(df_salesOrders_before_week_unwrapped['ui_status'] == 'Placed') & (df_salesOrders_before_week_unwrapped['ProductGroup'] == 'Bikes')]
    placed_bikes_before_week = df_placed_bikes_before_week['OrderNumber'].nunique()

   #accessores parts
    df_ready_to_ship_parts_accessories_before_week = (df_salesOrders_before_week_unwrapped.loc[(df_salesOrders_before_week_unwrapped['ui_status'] == 'Ready to Ship')
                                                              & (df_salesOrders_before_week_unwrapped['ProductGroup'].isin(get_parts_list() + ['Accessories']))])
    ready_to_ship_parts_accessories_before_week = df_ready_to_ship_parts_accessories_before_week['OrderNumber'].nunique()



    #during this week
    #bikes
    df_placed_bikes_week = df_salesOrders_week_unwrapped.loc[(df_salesOrders_week_unwrapped['ui_status'] == 'Placed') & (df_salesOrders_week_unwrapped['ProductGroup'] == 'Bikes')]
    placed_bikes_week = df_placed_bikes_week['OrderNumber'].nunique()

    #accessroes parts
    df_ready_to_ship_parts_accessories_week = (
    df_salesOrders_week_unwrapped.loc[(df_salesOrders_week_unwrapped['ui_status'] == 'Ready to Ship')
                                             & (df_salesOrders_week_unwrapped['ProductGroup'].isin(
        get_parts_list() + ['Accessories']))])
    ready_to_ship_parts_accessories_week = df_ready_to_ship_parts_accessories_week[
        'OrderNumber'].nunique()


    ops_summary_work_in_progres_ytd = pd.DataFrame({
        "Metric": [
            "Parked",
            "Backordered",
            "Hold + Accounting(Sales)",
            "Pre Orders",
            "Placed(Bikes) or Ready to Ship(parts & accessories)",
            "Total Open Orders",
            "# of Units associated with Pre-Orders"
        ],
        "Value": [
            old_parked + new_parked,
            old_backordered + new_backordered,
            old_hold + new_hold + old_accounting + new_accounting,
            "Zach",
            # placed_bikes_before_week + placed_bikes_week + ready_to_ship_parts_accessories_before_week + ready_to_ship_parts_accessories_week,
            old_placed + new_placed + old_ready_to_ship + new_ready_to_ship + old_ready_to_ship_STG + new_ready_to_ship_STG,
            old_open_orders + new_open_orders,
            "Zach"
        ]
    })
    print(ops_summary_work_in_progres_ytd)




    #BIKES
    df_salesOrders_week_unwrapped_Jamn_bikes = df_salesOrders_week_unwrapped.loc[(df_salesOrders_week_unwrapped['WarehouseName'] == 'Jam-N Logistics')
                                                                                 & (df_salesOrders_week_unwrapped['ProductGroup'] == 'Bikes')]
    Total_sent_Bike_Orders_JamN = df_salesOrders_week_unwrapped_Jamn_bikes['OrderNumber'].nunique()



    df_unwrapped_salesorders_completed_date = unwrap_warehouse_sales_orders(df_SalesOrders_completed_dates)
    df_unwrapped_salesorders_JamN_bikes_completed_dates = df_unwrapped_salesorders_completed_date.loc[
        (df_unwrapped_salesorders_completed_date['WarehouseName'] == 'Jam-N Logistics') &
        (df_unwrapped_salesorders_completed_date[
             'ProductGroup'] == 'Bikes')]
    Total_num_Bike_Orders_JamN = df_unwrapped_salesorders_JamN_bikes_completed_dates['OrderNumber'].nunique()
    num_bikes_fulfilled_JamN = df_unwrapped_salesorders_JamN_bikes_completed_dates['OrderQuantity'].sum()

    ops_bike_summary_completed = pd.DataFrame({
        "Metric": [
            "Total # of Bike Orders: Jam-N",
            "# of Bikes Fulfilled: Jam-N",
            "Jam-N Bikes Per Order"
        ],
        "Value": [
            Total_sent_Bike_Orders_JamN,
            num_bikes_fulfilled_JamN,
            num_bikes_fulfilled_JamN / Total_num_Bike_Orders_JamN
        ]
    })
    print(ops_bike_summary_completed)




    #Parts
    df_salesOrders_before_week_unwrapped = clean_sales_orders(df_salesOrders_before_week)
    df_salesOrders_week_unwrapped = clean_sales_orders(df_salesOrders_week)

    ready_to_ship_list = ['Ready to Ship','Ready Ship STG']
    df_parts_awaiting_shipment_before_week = df_salesOrders_before_week_unwrapped.loc[(df_salesOrders_before_week_unwrapped['ui_status'].isin(ready_to_ship_list))
                                                                                      & (df_salesOrders_before_week_unwrapped['ProductGroup'].isin(get_parts_list()))]
    total_parts_orders_orderDate_before_week = df_parts_awaiting_shipment_before_week['OrderNumber'].nunique()
    print("THING: ", df_parts_awaiting_shipment_before_week['OrderNumber'])

    df_parts_awaiting_shipment_week = df_salesOrders_week_unwrapped.loc[
        (df_salesOrders_week_unwrapped['ui_status'].isin(ready_to_ship_list))
        & (df_salesOrders_week_unwrapped['ProductGroup'].isin(get_parts_list()))]
    total_parts_orders_orderDate_week = df_parts_awaiting_shipment_week['OrderNumber'].nunique()
    print(df_parts_awaiting_shipment_week['OrderNumber'])



    print("///////////////////////")
    df_unwrapped_salesorders_parts_completed_dates = df_unwrapped_salesorders_completed_date.loc[df_unwrapped_salesorders_completed_date['ProductGroup'].isin(get_parts_list())]
    # df_unwrapped_salesorders_parts_completed_dates = df_unwrapped_salesorders_completed_date.loc[df_unwrapped_salesorders_completed_date['Warehouse'].isin(['Brickyard', ''])]
    # total_parts_orders = df_unwrapped_salesorders_parts_completed_dates['OrderNumber'].nunique()
    parts_units_fulfilled = df_unwrapped_salesorders_parts_completed_dates['OrderQuantity'].sum()

    df_unwrapped_salesorders_accessories_order_dates = df_unwrapped_salesorders_completed_date.loc[
        df_unwrapped_salesorders_completed_date['ProductGroup'] == 'Accessories']
    total_accessories_orders = df_unwrapped_salesorders_accessories_order_dates['OrderNumber'].nunique()
    accessory_units_fulfilled = df_unwrapped_salesorders_accessories_order_dates['OrderQuantity'].sum()

    ops_parts_accessories_summary_completed = pd.DataFrame({
        "Metric": [
            "Total Parts Orders Completed",
            "Parts Units Fulfilled",
            "Total Accessories Orders",
            "Accessory Units Fulfilled"
        ],
        "Value": [
            total_parts_orders_orderDate_before_week + total_parts_orders_orderDate_week,
            parts_units_fulfilled,
            total_accessories_orders,
            accessory_units_fulfilled
        ]
    })
    print(ops_parts_accessories_summary_completed)









    # MAKE EXCEL FILE
    from openpyxl import load_workbook
    from openpyxl.styles import Font


    # Collect all summaries with titles
    sections = [
        ("Productivity WTD", ops_summary_productivity_WTD),
        ("Work in Progress YTD", ops_summary_work_in_progres_ytd),
        ("Bike Orders", ops_bike_summary_completed),
        ("Parts Orders", ops_parts_accessories_summary_completed)
    ]

    # Build a combined DataFrame with titles + spacing
    combined = []
    for title, df in sections:
        # Title row
        combined.append(pd.DataFrame({"Metric": [title], "Value": [""]}))
        # Actual section
        combined.append(df)
        # Blank spacer row
        combined.append(pd.DataFrame({"Metric": [""], "Value": [""]}))

    final_df = pd.concat(combined, ignore_index=True)

    # Save to Excel at the requested path
    file_path = fr"C:\Users\joshu\Documents\Ops_Dashboard_Draft2.xlsx"
    final_df.to_excel(file_path, index=False)

    # Optional: format titles bold
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):  # only "Metric" column
        cell = row[0]
        if cell.value in [title for title, _ in sections]:
            cell.font = Font(bold=True)
    wb.save(file_path)







# first_try_dashboard()
third_try_dashboard()